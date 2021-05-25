from tkinter import *
import tkinter as tk
import datetime
import self as self
import tk
import yfinance as yf
import openpyxl as xl
import xlsxwriter


# function returns current price of stock
def get_current_price(symbol):
    ticker = yf.Ticker(symbol)
    todays_data = ticker.history(period='1d')
    return todays_data['Close'][0]

#date & time for parsing 
today = datetime.datetime.now()
d = datetime.timedelta(days = 30)
a = today - d
#print(today.strftime('%Y-%m-%d'))
print(a)


# testing with yfinance package
msft = yf.Ticker("MSFT")
# str = msft.info
# hist = msft.history(period="max")
rec = msft.recommendations
reccc = str(rec)
reccc.index(str(a))
print(reccc)
#print(rec)
#test example
recList = []
q = 0
tickerList = ["AAPL", "MSFT", "GOOGL", "ABNB", "AMZN", "SPY"]


# for tick in tickerList:
# stockName = yf.Ticker(tickerList[q])
# recList.append(stockName.recommendations)
# q += 1
# print(recList[1])
# function to get a list of prices
def listPrice(tickerList):
    priceList = []
    for x in tickerList:
        priceList.append(("${:,.2f}".format(get_current_price(x))))
    return priceList


print(listPrice(tickerList))

# function to get reccomendations
# Excel management
# def process_workbook():
# sPrice = mysheet.cell(row=4, column=5)
# wb = xl.load_workbook(filename='stock.xlsx')
# sheet = wb['Sheet1']
# cell = sheet['a1']
# return cell.value()

workbook = xlsxwriter.Workbook('~$stock.xlsx')
worksheet = workbook.add_worksheet()
for col_num, data in enumerate(listPrice(tickerList)):
    worksheet.write(1, col_num + 1, data)

for col_num, data in enumerate(tickerList):
    worksheet.write(0, col_num + 1, data)

# for col_num, data in enumerate(tickerList):
# worksheet.write(2, col_num + 1, data)

workbook.close()
# for x in tickerList:


# UI
# app = Tk() # the application itself
# app.title("yfinance tester") # title of window

# label = Label(app, text="") # creates label
# tagline = Label(app, text = str)
# label.pack() # adds the label to the window
# tagline.pack()


# Returns val to UI
# canvas1 = tk.Canvas(app, width = 400, height = 300)
# canvas1.pack()
# entry1 = tk.Entry(app)
# canvas1.create_window(200, 140, window=entry1)

# def getPrice():
#   x1 = entry1.get()

#  label1 = tk.Label(app, text=get_current_price(x1))
# canvas1.create_window(200, 230, window=label1)
# button1 = tk.Button(text='Get the Price', command=getPrice())
# canvas1.create_window(200, 180, window=button1)


# app.mainloop()
